import pandas as pd
from sqlalchemy import create_engine
import openpyxl
import time
import re
import os
import glob
import time
from datetime import datetime

def escape_special_chars(value):
    """Escape special characters for SQL query"""
    if value is None:
        return "NULL"
    return str(value).replace("'", "''")  # Escape single quotes

class SQLQueryRunner:
    def __init__(self, server, database, username, password, port=1433):
        self.server = server
        self.database = database
        self.username = username
        self.password = password
        self.port = port
        self.engine = None

    def connect_to_sql(self):
        """Establish SQL Server connection using SQLAlchemy"""
        try:
            connection_string = f"mssql+pytds://{self.username}:{self.password}@{self.server}:{self.port}/{self.database}"
            self.engine = create_engine(connection_string)
            # Test the connection
            with self.engine.connect() as conn:
                pass
            print("✅ Connected to SQL Server successfully!")
            return True
        except Exception as e:
            print(f"❌ Error connecting to SQL Server: {e}")
            return False

    def execute_query(self, query):
        """Execute a SQL query and return a DataFrame"""
        if not self.engine:
            if not self.connect_to_sql():
                return None
        try:
            return pd.read_sql(query, self.engine)
        except Exception as e:
            print(f"❌ Error executing query: {e}")
            return None


def read_excel_with_hidden_columns(file_path):
    """Read Excel file including hidden columns and preserve column visibility info"""
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)

    # Get the active sheet (or you can specify a sheet by name)
    sheet = workbook.active

    # Store column visibility information
    column_visibility = {}
    for col_letter, col_dimension in sheet.column_dimensions.items():
        column_visibility[col_letter] = {
            'hidden': col_dimension.hidden,
            'width': col_dimension.width
        }

    # Print hidden columns info
    hidden_cols = [col for col, info in column_visibility.items() if info['hidden']]
    print(f"Found {len(hidden_cols)} hidden columns: {', '.join(hidden_cols)}")

    # Read all data including hidden columns
    all_data = []
    for row in sheet.iter_rows(values_only=True):
        all_data.append(row)

    # Convert to DataFrame
    df = pd.DataFrame(all_data)

    # Handle headers (first row)
    headers = df.iloc[0]
    df = df[1:]
    df.columns = headers

    return df, column_visibility, workbook


def copy_excel_with_exact_format(input_file, output_file):
    """Create an exact copy of the Excel file with all formatting preserved"""
    # Load the workbook
    wb = openpyxl.load_workbook(input_file)
    # Save as new file
    wb.save(output_file)
    return wb

def is_default_date(date_str):
    """Identify a default date (01/01/1900) regardless of format"""
    if date_str is None or date_str == "":
        return False

    # Convert to string if it's not already
    date_str = str(date_str).strip()

    # Try specific string matches first for 1900
    if date_str == '01/01/1900' or date_str == '1900-01-01' or date_str == '1900/01/01':
        return True

    # Check if it contains the string pattern for 1900
    if re.search(r'01.01.1900|1900.01.01', date_str.replace('/', '.').replace('-', '.')):
        return True

    # Try to parse as a date object and compare
    try:
        # Try multiple formats
        for fmt in ['%m/%d/%Y', '%Y-%m-%d', '%m-%d-%Y', '%Y/%m/%d']:
            try:
                date_obj = datetime.strptime(date_str, fmt)
                if date_obj.year == 1900 and date_obj.month == 1 and date_obj.day == 1:
                    return True
            except ValueError:
                continue
    except Exception as e:
        print(f"Error parsing date '{date_str}': {e}")

    return False

def parse_date_string(date_str):
    """Parse date string and return datetime object or None"""
    if not date_str or date_str == "":
        return None

    date_str = str(date_str).strip()

    # Try multiple date formats
    formats = ['%m/%d/%Y', '%Y-%m-%d', '%m-%d-%Y', '%Y/%m/%d', '%d/%m/%Y']

    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue

    return None

def format_date_for_comment(date_str):
    """Format date for comment field as dd/mm/yyyy"""
    if not date_str or date_str == "":
        return ""

    date_obj = parse_date_string(date_str)
    if date_obj:
        return date_obj.strftime("%d/%m/%Y")
    return str(date_str)  # Return original string if parsing fails

def format_date_for_inspection(date_str):
    """Format date for factory inspection plan as dd/mm (day and month only)"""
    if not date_str or date_str == "":
        return ""

    date_obj = parse_date_string(date_str)
    if date_obj:
        return date_obj.strftime("%d/%m")
    return str(date_str)  # Return original string if parsing fails

def construct_query(material, po):
    """Properly construct a SQL query with proper escaping to avoid formatting issues"""
    # Sanitize inputs
    material_safe = str(material).replace("'", "''") if material else ""
    po_safe = str(po).replace("'", "''") if po else ""

    # Manually build the where clause with proper escaping
    where_parts = []

    if material_safe:
        where_parts.append(f"scp.cust_ucd = '{material_safe}'")

    if po_safe:
        po_part = f"((om.cust_po = '{po_safe}' OR om.cust_po LIKE '%%{po_safe}%%') AND om.customer != 'MSE-INTERNAL')"

        numeric_parts = ''.join(c for c in po_safe if c.isdigit())
        if numeric_parts:
            po_part = f"({po_part} OR (om.cust_po LIKE '%%{numeric_parts}%%' AND om.customer != 'MSE-INTERNAL'))"

        where_parts.append(po_part)

    # Default condition if no filters provided
    if not where_parts:
        where_parts = ["1=1"]

    # Join with AND
    where_clause = " AND ".join(where_parts)

    # Hardcode the SQL template right here to avoid any issues with format()
    # UPDATED: Check for 01/01/1900 dates directly in the query
    # Modified to remove time part by using CONVERT with style 101 (MM/DD/YYYY)
    query_template = """
        SELECT
            RTRIM(ot.article_no) + '/' + CAST(ot.uniq_code AS VARCHAR) AS [Sterling Code/Unique Code],
            om.fac_order AS [Factory Order No],
            CONVERT(VARCHAR, om.sched1_dt, 101) AS [Factory Order Date],
            CASE WHEN CONVERT(VARCHAR, om.lbl_dt, 101) = '01/01/1900' THEN '' 
                 ELSE CONVERT(VARCHAR, om.lbl_dt, 101) END AS [Label Date],
            CONVERT(VARCHAR, om.rev_cdate, 101) AS [Customer Date]
        FROM order_trn ot
        JOIN sp_cust_price scp 
            ON ot.article_no = scp.art_code AND ot.uniq_code = scp.uniq_code
        JOIN order_mast om
            ON ot.fac_order = om.fac_order
        WHERE """

    # Directly concatenate the where_clause instead of using format()
    final_query = query_template + where_clause

    return final_query

def construct_invoice_query(factory_order_no, article_no, unique_no):
    """Properly construct an invoice query with proper escaping to avoid formatting issues"""
    if not factory_order_no or str(factory_order_no).strip() == "":
        return None

    # Sanitize inputs
    fo_safe = str(factory_order_no).strip().replace("'", "''")
    article_safe = str(article_no).strip().replace("'", "''") if article_no and str(article_no).strip() != "" else ""
    unique_safe = str(unique_no).strip().replace("'", "''") if unique_no and str(unique_no).strip() != "" else ""

    # Manually build the where clause
    where_parts = [f"p.ord = '{fo_safe}'"]

    if article_safe:
        where_parts.append(f"ot.article_no = '{article_safe}'")

    if unique_safe:
        where_parts.append(f"ot.uniq_code = '{unique_safe}'")

    # Join with AND
    where_clause = " AND ".join(where_parts)

    # Hardcode the SQL template here to avoid issues with format()
    # Modified to convert date without time using CONVERT with style 101 (MM/DD/YYYY)
    query_template = """
    SELECT 
        em.inv_no AS [Invoice Number], 
        CONVERT(VARCHAR, em.inv_date, 101) AS [Invoice Date],
        CAST(p.pcs AS VARCHAR) AS [Shipped Quantity]
    FROM order_mast o 
    INNER JOIN order_Trn ot ON o.fac_order = ot.fac_order
    INNER JOIN pkt_trns p ON ot.fac_order = p.ord AND ot.article_no = p.art AND ot.uniq_code = p.uc
    INNER JOIN export_mast em ON p.inv_no = em.inv_no 
    WHERE """

    # Directly concatenate the where_clause instead of using format()
    final_query = query_template + where_clause

    return final_query

def update_comment_and_inspection_columns(output_file):
    """Update Comment and Factory Inspection Plan columns based on Customer Date"""
    print("\nUpdating Comment and Factory Inspection Plan columns...")

    # Load the output workbook
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Find header row and column indexes
    header_row = 1
    customer_date_col_idx = None
    comment_col_idx = None
    inspection_col_idx = None

    # Find column indexes
    for col_idx, cell in enumerate(ws[header_row], start=1):
        cell_value = str(cell.value).strip() if cell.value else ""
        if cell_value == 'Customer Date':
            customer_date_col_idx = col_idx
        elif cell_value == 'Comment':
            comment_col_idx = col_idx
        elif cell_value == 'Factory Inspection Plan':
            inspection_col_idx = col_idx

    # Check if Customer Date column exists
    if customer_date_col_idx is None:
        print("⚠️ Customer Date column not found - cannot update Comment and Factory Inspection Plan")
        wb.close()
        return

    # Add Comment column if it doesn't exist
    if comment_col_idx is None:
        max_col = ws.max_column
        comment_col_idx = max_col + 1
        ws.cell(row=header_row, column=comment_col_idx).value = 'Comment'
        print(f"Added 'Comment' column at position {comment_col_idx}")

    # Add Factory Inspection Plan column if it doesn't exist
    if inspection_col_idx is None:
        max_col = ws.max_column
        inspection_col_idx = max_col + 1
        ws.cell(row=header_row, column=inspection_col_idx).value = 'Factory Inspection Plan'
        print(f"Added 'Factory Inspection Plan' column at position {inspection_col_idx}")

    # Process each row
    comment_updates = 0
    inspection_updates = 0

    for row_idx in range(header_row + 1, ws.max_row + 1):
        # Get Customer Date value
        customer_date_value = ws.cell(row=row_idx, column=customer_date_col_idx).value

        if customer_date_value and str(customer_date_value).strip() != "":
            customer_date_str = str(customer_date_value).strip()

            # Update Comment column if empty
            comment_cell = ws.cell(row=row_idx, column=comment_col_idx)
            if not comment_cell.value or str(comment_cell.value).strip() == "":
                formatted_date = format_date_for_comment(customer_date_str)
                if formatted_date:
                    comment_text = f"This item is schedule to be ready by {formatted_date}"
                    comment_cell.value = comment_text
                    comment_updates += 1

            # Update Factory Inspection Plan column if empty
            inspection_cell = ws.cell(row=row_idx, column=inspection_col_idx)
            if not inspection_cell.value or str(inspection_cell.value).strip() == "":
                formatted_inspection_date = format_date_for_inspection(customer_date_str)
                if formatted_inspection_date:
                    inspection_cell.value = formatted_inspection_date
                    inspection_updates += 1

    # Save the workbook
    wb.save(output_file)
    wb.close()

    print(f"✅ Updated {comment_updates} Comment fields")
    print(f"✅ Updated {inspection_updates} Factory Inspection Plan fields")

def update_excel_with_query_results(output_file, material_column, po_column, query_results):
    """Update the Excel file with SQL query results while preserving format"""
    # Load the output workbook
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Find header row and column indexes
    header_row = 1  # Assuming first row contains headers
    material_col_idx = None
    po_col_idx = None

    # Get column indexes from headers
    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value == material_column:
            material_col_idx = col_idx
        elif cell.value == po_column:
            po_col_idx = col_idx

    if material_col_idx is None or po_col_idx is None:
        print(f"⚠️ Could not find Material or PO column in the Excel file")
        print(f"Looking for '{material_column}' and '{po_column}'")
        print(f"Available headers: {[cell.value for cell in ws[header_row]]}")
        return [], None

    # Add new columns for SQL results if they don't exist
    new_columns = ['Sterling Code/Unique Code', 'Factory Order No', 'Factory Order Date',
                   'Label Date', 'Customer Date']

    # Find the last column index
    max_col = ws.max_column
    col_mappings = {}

    # Check if columns already exist, otherwise add them
    for col_name in new_columns:
        col_exists = False
        for col_idx in range(1, max_col + 1):
            if ws.cell(row=header_row, column=col_idx).value == col_name:
                col_exists = True
                col_mappings[col_name] = col_idx
                break

        if not col_exists:
            max_col += 1
            ws.cell(row=header_row, column=max_col).value = col_name
            col_mappings[col_name] = max_col
            print(f"Added new column '{col_name}' at position {max_col}")

    # Iterate through data rows and update with query results
    rows_updated = 0
    total_rows = 0
    rows_with_data = 0
    factory_order_data = []  # Store factory orders with article numbers and unique codes
    factory_order_col_idx = col_mappings.get('Factory Order No', None)
    sterling_code_col_idx = col_mappings.get('Sterling Code/Unique Code', None)
    label_date_col_idx = col_mappings.get('Label Date', None)

    for row_idx in range(header_row + 1, ws.max_row + 1):
        total_rows += 1
        # Get Material and PO values from the row
        material_value = ws.cell(row=row_idx, column=material_col_idx).value
        po_value = ws.cell(row=row_idx, column=po_col_idx).value

        # Skip completely empty rows
        if material_value is None and po_value is None:
            continue

        rows_with_data += 1

        # Standardize material and PO values exactly like when building the result_map
        material_str = str(material_value).strip() if material_value is not None else ""

        # Handle PO values preserving their original format
        if pd.isna(po_value) or po_value is None:
            po_str = ""
        else:
            po_str = str(po_value).strip()

        # Debug print some rows to see what's happening
        if rows_with_data <= 5 or rows_with_data % 100 == 0:
            print(f"Row {row_idx}: Material='{material_str}', PO='{po_str}'")

        # Check if we have results for this Material/PO combination
        key = (material_str, po_str)

        if key in query_results and not query_results[key].empty:
            # Get the first row of results for this key
            result_data = query_results[key].iloc[0].to_dict()

            # Handle Label Date - replace 01/01/1900 with empty string
            if 'Label Date' in result_data and is_default_date(result_data['Label Date']):
                result_data['Label Date'] = ''

            # Update cells with result data
            for col_name, col_idx in col_mappings.items():
                if col_name in result_data:
                    # Double-check Label Date again right before writing to Excel
                    if col_name == 'Label Date' and is_default_date(result_data[col_name]):
                        ws.cell(row=row_idx, column=col_idx).value = ''
                    else:
                        ws.cell(row=row_idx, column=col_idx).value = result_data[col_name]

            # If this is the Factory Order No column and Sterling Code column, store their values
            factory_order_value = result_data.get('Factory Order No')
            sterling_code = result_data.get('Sterling Code/Unique Code')

            # We need article_no and unique_no for invoice query, so extract them from Sterling Code
            if sterling_code and '/' in sterling_code:
                article_no = sterling_code.split('/')[0]
                unique_no = sterling_code.split('/')[1]
            else:
                article_no = ""
                unique_no = ""

            if factory_order_value is not None:
                factory_order_data.append((row_idx, factory_order_value, article_no, unique_no))

            rows_updated += 1

    # ADDITIONAL LABEL DATE SCAN: Check the entire sheet for any missed instances
    if label_date_col_idx:
        print("\nPerforming additional Label Date check across entire worksheet...")
        fixed_count = 0

        for row_idx in range(header_row + 1, ws.max_row + 1):
            label_date_value = ws.cell(row=row_idx, column=label_date_col_idx).value

            # Handle any variant of the 01/01/1900 date
            if label_date_value is not None and is_default_date(label_date_value):
                ws.cell(row=row_idx, column=label_date_col_idx).value = ''
                fixed_count += 1

        if fixed_count > 0:
            print(f"Fixed an additional {fixed_count} rows with default date (01/01/1900) values")

    # Save the updated workbook
    wb.save(output_file)
    wb.close()  # Explicitly close the workbook

    print(f"Processed {total_rows} total rows, {rows_with_data} rows with data")
    print(f"Updated {rows_updated} rows in {output_file}")

    return factory_order_data, factory_order_col_idx

def update_excel_with_invoice_data(output_file, factory_order_data, factory_order_col_idx, invoice_results):
    """Update the Excel file with invoice query results"""
    # Load the output workbook
    wb = openpyxl.load_workbook(output_file)
    ws = wb.active

    # Find header row
    header_row = 1

    # Add new columns for invoice data if they don't exist
    invoice_columns = ['Invoice Number', 'Invoice Date', 'Shipped Quantity']

    # Find the last column index
    max_col = ws.max_column
    inv_col_mappings = {}

    # FIRST ENSURE ALL COLUMNS EXIST IN THE EXCEL FILE
    for col_name in invoice_columns:
        col_exists = False
        for col_idx in range(1, max_col + 1):
            if ws.cell(row=header_row, column=col_idx).value == col_name:
                col_exists = True
                inv_col_mappings[col_name] = col_idx
                print(f"Found existing column '{col_name}' at position {col_idx}")
                break

        if not col_exists:
            max_col += 1
            ws.cell(row=header_row, column=max_col).value = col_name
            inv_col_mappings[col_name] = max_col
            print(f"Added new column '{col_name}' at position {max_col}")

    # Double verify that the Shipped Qty column exists
    if 'Shipped Quantity' not in inv_col_mappings:
        print(f"⚠️ CRITICAL: 'Shipped Quantity' column was not added properly!")
        # Force add it again
        max_col += 1
        ws.cell(row=header_row, column=max_col).value = 'Shipped Quantity'
        inv_col_mappings['Shipped Quantity'] = max_col
        print(f"🔄 Forcibly added 'Shipped Quantity' column at position {max_col}")

    # Track number of rows updated with invoice data
    invoice_rows_updated = 0

    # VERIFY AND FIX LABEL DATE VALUES ACROSS THE ENTIRE SHEET
    # Check for Label Date column
    label_date_col_idx = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=col_idx).value == 'Label Date':
            label_date_col_idx = col_idx
            break

    # Fix any Label Date values that are still showing as 01/01/1900
    if label_date_col_idx:
        print("Scanning entire sheet for any Label Date = 01/01/1900 values...")
        fixed_count = 0
        for row_idx in range(header_row + 1, ws.max_row + 1):
            label_date_value = ws.cell(row=row_idx, column=label_date_col_idx).value
            # Check for the exact string '01/01/1900'
            if label_date_value == '01/01/1900':
                ws.cell(row=row_idx, column=label_date_col_idx).value = ''
                fixed_count += 1
            # Also use the function for more complex cases
            elif label_date_value and is_default_date(label_date_value):
                ws.cell(row=row_idx, column=label_date_col_idx).value = ''
                fixed_count += 1

        if fixed_count > 0:
            print(f"Fixed {fixed_count} rows with default Label Date values")

    # Iterate through factory order numbers and update with invoice data
    for row_idx, factory_order_no, article_no, unique_no in factory_order_data:
        # Skip if factory order number is None or empty
        if factory_order_no is None or str(factory_order_no).strip() == "":
            continue

        # Normalize factory order number
        fo_str = str(factory_order_no).strip()

        # Check if we have invoice results for this factory order number
        if fo_str in invoice_results and not invoice_results[fo_str].empty:
            # We might have multiple invoices per factory order - get the first one for now
            invoice_data = invoice_results[fo_str].iloc[0].to_dict()

            # Update cells with invoice data
            for col_name, col_idx in inv_col_mappings.items():
                if col_name in invoice_data:
                    # For Shipped Qty specifically, ensure it's a string and force writing
                    if col_name == 'Shipped Quantity':
                        value = str(invoice_data[col_name]) if invoice_data[col_name] is not None else ""
                        ws.cell(row=row_idx, column=col_idx).value = value
                    else:
                        value = invoice_data[col_name]
                        ws.cell(row=row_idx, column=col_idx).value = value

            invoice_rows_updated += 1

    # Make sure all columns are really there by adding them again if needed
    for col_name in invoice_columns:
        found = False
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=col_idx).value == col_name:
                found = True
                break

        if not found:
            print(f"⚠️ WARNING: Column '{col_name}' is still missing! Adding it one more time.")
            ws.cell(row=header_row, column=ws.max_column + 1).value = col_name

    # Save the updated workbook
    wb.save(output_file)
    wb.close()  # Explicitly close the workbook
    print(f"Updated {invoice_rows_updated} rows with invoice data in {output_file}")

    # Return success status based on whether Shipped Qty column was added
    return 'Shipped Quantity' in inv_col_mappings

def get_recent_files(folder_path, file_pattern="*.xlsx", max_files=5):
    """
    Get the most recently added/modified files in the specified folder

    Args:
        folder_path: Path to the folder to search
        file_pattern: Pattern to match files (e.g., "*.xlsx")
        max_files: Maximum number of files to return

    Returns:
        List of file paths sorted by modification time (newest first)
    """
    # Make sure folder path exists
    if not os.path.exists(folder_path):
        print(f"⚠️ Folder does not exist: {folder_path}")
        return []

    # Get all files matching the pattern
    files = glob.glob(os.path.join(folder_path, file_pattern))

    if not files:
        print(f"⚠️ No {file_pattern} files found in {folder_path}")
        return []

    # Sort files by modification time (newest first)
    files_with_time = [(f, os.path.getmtime(f)) for f in files]
    sorted_files = sorted(files_with_time, key=lambda x: x[1], reverse=True)

    # Return the paths of the most recent files
    recent_files = [f[0] for f in sorted_files[:max_files]]

    print(f"✅ Found {len(recent_files)} recent {file_pattern} files in {folder_path}")
    for i, file in enumerate(recent_files):
        print(f"  {i+1}. {os.path.basename(file)} - Modified: {time.ctime(os.path.getmtime(file))}")

    return recent_files

def generate_output_filename(input_file):
    """
    Generate an output filename by adding *Auto*_Generated to the filename

    Args:
        input_file: Original filename

    Returns:
        New filename with *Auto*_Generated added
    """
    # Split the path, filename, and extension
    dir_path = os.path.dirname(input_file)
    base_name = os.path.basename(input_file)
    name, ext = os.path.splitext(base_name)

    # Create new filename with *Auto*_Generated
    new_name = f"{name}_*Auto*_Generated{ext}"

    # If directory path exists, join it with the new filename
    if dir_path:
        new_path = os.path.join(dir_path, new_name)
    else:
        new_path = new_name

    return new_path


# Main execution function
def process_file(input_file, output_file=None):
    """
    Process a single file with the SQL queries and updates

    Args:
        input_file: Path to the input Excel file
        output_file: Path to the output Excel file (if None, will be generated)

    Returns:
        Path to the output file
    """
    # Start timer
    start_time = time.time()

    # Generate output filename if not provided
    if output_file is None:
        output_file = generate_output_filename(input_file)

    print(f"\n{'='*80}")
    print(f"PROCESSING FILE: {input_file}")
    print(f"OUTPUT FILE: {output_file}")
    print(f"{'='*80}\n")

    # --- Configuration ---
    server = "111.93.56.76"
    database = "SOPL"
    username = "Krutika"
    password = "K123456"

    # Show all columns in the output
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)
    pd.set_option('display.max_colwidth', None)

    # --- Instantiate SQL runner ---
    runner = SQLQueryRunner(server, database, username, password)

    # --- Read Excel file including hidden columns ---
    print(f"Reading Excel file: {input_file}")
    excel_data, column_visibility, original_workbook = read_excel_with_hidden_columns(input_file)

    # Find material (column H) and PO (column D)
    material_column = 'Material'  # Adjust if your actual column name is different
    po_column = 'PO'  # Adjust if your actual column name is different

    # Check if columns exist by name
    if material_column not in excel_data.columns:
        # Try by index (column H is typically index 7)
        material_column = excel_data.columns[7]
        print(f"Using column {material_column} for Material")

    if po_column not in excel_data.columns:
        # Try by index (column D is typically index 3)
        po_column = excel_data.columns[3]
        print(f"Using column {po_column} for PO")

    # Process all rows even with some null values (as long as we have either Material or PO)
    material_po_pairs = excel_data[[material_column, po_column]].copy()
    material_po_pairs = material_po_pairs[~(material_po_pairs[material_column].isna() & material_po_pairs[po_column].isna())]
    material_po_pairs = material_po_pairs.drop_duplicates()

    print(f"Found {len(material_po_pairs)} Material/PO combinations to process")

    # Create an exact copy of the input file as our output file
    print(f"Creating output file: {output_file}")
    copy_excel_with_exact_format(input_file, output_file)

    # ---- PHASE 1: Get Factory Order details ----
    print("PHASE 1: Executing SQL queries for Factory Orders...")
    query_results = {}
    total_results_rows = 0

    # Process in batches
    for idx, row in material_po_pairs.iterrows():
        # Get material and PO values
        material = str(row[material_column]).strip() if not pd.isna(row[material_column]) else ""

        # Handle PO values without trying to convert to integer
        po = str(row[po_column]).strip() if not pd.isna(row[po_column]) else ""

        # Skip if both are empty
        if not material and not po:
            continue

        # Print info about the query
        print(f"Querying Material: '{material}', PO: '{po}'")

        # Execute the query
        query = construct_query(material, po)
        result = runner.execute_query(query)

        key = (material, po)
        if result is not None and not result.empty:
            query_results[key] = result
            total_results_rows += len(result)
            print(f"  ✓ Found {len(result)} results")
        else:
            query_results[key] = pd.DataFrame()  # Empty DataFrame for no results
            print("  × No results found")

    print(f"\nRetrieved {total_results_rows} total rows from SQL queries")
    print(f"Found data for {len([k for k, v in query_results.items() if not v.empty])}/{len(query_results)} Material/PO combinations")

    # Now update the Excel file with all the query results
    print("\nUpdating Excel file with Factory Order results...")
    factory_order_data, factory_order_col_idx = update_excel_with_query_results(output_file, material_column, po_column, query_results)

    # Add a small delay to ensure file is completely saved
    time.sleep(2)

    # ---- PHASE 2: Get Invoice details ----
    print(f"\nPHASE 2: Executing SQL queries for Invoice Data...")
    print(f"Found {len(factory_order_data)} Factory Order numbers to query for invoices")

    # Execute invoice queries and store results
    invoice_results = {}

    for i, (row_idx, factory_order_no, article_no, unique_no) in enumerate(factory_order_data):
        # Skip if factory order number is None or empty
        if factory_order_no is None or str(factory_order_no).strip() == "":
            continue

        # Convert to string and standardize format
        fo_str = str(factory_order_no).strip()

        # Only display progress for some orders to avoid cluttering the output
        if i < 5 or i % 50 == 0:
            print(f"Querying Invoice data for Factory Order: '{fo_str}', Article No: '{article_no}', Unique Code: '{unique_no}' (Order {i+1}/{len(factory_order_data)})")

        # Execute the invoice query with article_no and unique_no
        query = construct_invoice_query(fo_str, article_no, unique_no)
        if query:
            result = runner.execute_query(query)

            if result is not None and not result.empty:
                invoice_results[fo_str] = result
                if i < 5 or i % 50 == 0:
                    print(f"  ✓ Found {len(result)} invoices")
            else:
                invoice_results[fo_str] = pd.DataFrame()  # Empty DataFrame for no results
                if i < 5 or i % 50 == 0:
                    print("  × No invoices found")

    print(f"\nRetrieved invoice data for {len([k for k, v in invoice_results.items() if not v.empty])}/{len(factory_order_data)} Factory Orders")

    # Update Excel file with invoice data
    print("\nUpdating Excel file with Invoice results...")
    shipped_qty_added = update_excel_with_invoice_data(output_file, factory_order_data, factory_order_col_idx, invoice_results)

    # ---- PHASE 3: Update Comment and Factory Inspection Plan columns ----
    print(f"\nPHASE 3: Updating Comment and Factory Inspection Plan columns...")
    update_comment_and_inspection_columns(output_file)

    # Calculate processing time
    end_time = time.time()
    duration = end_time - start_time

    print(f"\n{'='*80}")
    print(f"✅ PROCESSING COMPLETED SUCCESSFULLY!")
    print(f"📁 Output file: {output_file}")
    print(f"⏱️ Total processing time: {duration:.2f} seconds")
    print(f"{'='*80}")

    return output_file

# Alternative: Modify the existing main function to have a simple_return parameter
def main_with_option(input_file, output_file=None, server="111.93.56.76", database="SOPL",
                    username="Krutika", password="K123456", simple_return=False):
    """
    Main function with option to return just the file path or full details

    Args:
        input_file: Path to the input Excel file
        output_file: Path to the output Excel file (if None, will be generated)
        server: SQL Server address
        database: Database name
        username: Database username
        password: Database password
        simple_return: If True, returns only output file path; if False, returns full dict

    Returns:
        str or dict: Output file path (if simple_return=True) or full results dict
    """
    try:
        # Start timer
        start_time = time.time()

        # Generate output filename if not provided
        if output_file is None:
            output_file = generate_output_filename(input_file)

        print(f"\n{'='*80}")
        print(f"PROCESSING FILE: {input_file}")
        print(f"OUTPUT FILE: {output_file}")
        print(f"{'='*80}\n")

        # Show all columns in the output
        pd.set_option('display.max_columns', None)
        pd.set_option('display.width', None)
        pd.set_option('display.max_colwidth', None)

        # --- Instantiate SQL runner ---
        runner = SQLQueryRunner(server, database, username, password)

        # --- Read Excel file including hidden columns ---
        print(f"Reading Excel file: {input_file}")
        excel_data, column_visibility, original_workbook = read_excel_with_hidden_columns(input_file)

        # Find material (column H) and PO (column D)
        material_column = 'Material'
        po_column = 'PO'

        # Check if columns exist by name
        if material_column not in excel_data.columns:
            # Try by index (column H is typically index 7)
            material_column = excel_data.columns[7]
            print(f"Using column {material_column} for Material")

        if po_column not in excel_data.columns:
            # Try by index (column D is typically index 3)
            po_column = excel_data.columns[3]
            print(f"Using column {po_column} for PO")

        # Process all rows even with some null values
        material_po_pairs = excel_data[[material_column, po_column]].copy()
        material_po_pairs = material_po_pairs[~(material_po_pairs[material_column].isna() & material_po_pairs[po_column].isna())]
        material_po_pairs = material_po_pairs.drop_duplicates()

        print(f"Found {len(material_po_pairs)} Material/PO combinations to process")

        # Create an exact copy of the input file as our output file
        print(f"Creating output file: {output_file}")
        copy_excel_with_exact_format(input_file, output_file)

        # ---- PHASE 1: Get Factory Order details ----
        print("PHASE 1: Executing SQL queries for Factory Orders...")
        query_results = {}
        total_results_rows = 0

        # Process in batches
        for idx, row in material_po_pairs.iterrows():
            # Get material and PO values
            material = str(row[material_column]).strip() if not pd.isna(row[material_column]) else ""
            po = str(row[po_column]).strip() if not pd.isna(row[po_column]) else ""

            # Skip if both are empty
            if not material and not po:
                continue

            print(f"Querying Material: '{material}', PO: '{po}'")

            # Execute the query
            query = construct_query(material, po)
            result = runner.execute_query(query)

            key = (material, po)
            if result is not None and not result.empty:
                query_results[key] = result
                total_results_rows += len(result)
                print(f"  ✓ Found {len(result)} results")
            else:
                query_results[key] = pd.DataFrame()
                print("  × No results found")

        print(f"\nRetrieved {total_results_rows} total rows from SQL queries")
        print(f"Found data for {len([k for k, v in query_results.items() if not v.empty])}/{len(query_results)} Material/PO combinations")

        # Update the Excel file with all the query results
        print("\nUpdating Excel file with Factory Order results...")
        factory_order_data, factory_order_col_idx = update_excel_with_query_results(output_file, material_column, po_column, query_results)

        # Add a small delay to ensure file is completely saved
        time.sleep(2)

        # ---- PHASE 2: Get Invoice details ----
        print(f"\nPHASE 2: Executing SQL queries for Invoice Data...")
        print(f"Found {len(factory_order_data)} Factory Order numbers to query for invoices")

        # Execute invoice queries and store results
        invoice_results = {}

        for i, (row_idx, factory_order_no, article_no, unique_no) in enumerate(factory_order_data):
            # Skip if factory order number is None or empty
            if factory_order_no is None or str(factory_order_no).strip() == "":
                continue

            # Convert to string and standardize format
            fo_str = str(factory_order_no).strip()

            # Only display progress for some orders to avoid cluttering the output
            if i < 5 or i % 50 == 0:
                print(f"Querying Invoice data for Factory Order: '{fo_str}', Article No: '{article_no}', Unique Code: '{unique_no}' (Order {i+1}/{len(factory_order_data)})")

            # Execute the invoice query with article_no and unique_no
            query = construct_invoice_query(fo_str, article_no, unique_no)
            if query:
                result = runner.execute_query(query)

                if result is not None and not result.empty:
                    invoice_results[fo_str] = result
                    if i < 5 or i % 50 == 0:
                        print(f"  ✓ Found {len(result)} invoices")
                else:
                    invoice_results[fo_str] = pd.DataFrame()
                    if i < 5 or i % 50 == 0:
                        print("  × No invoices found")

        print(f"\nRetrieved invoice data for {len([k for k, v in invoice_results.items() if not v.empty])}/{len(factory_order_data)} Factory Orders")

        # Update Excel file with invoice data
        print("\nUpdating Excel file with Invoice results...")
        shipped_qty_added = update_excel_with_invoice_data(output_file, factory_order_data, factory_order_col_idx, invoice_results)

        # ---- PHASE 3: Update Comment and Factory Inspection Plan columns ----
        print(f"\nPHASE 3: Updating Comment and Factory Inspection Plan columns...")
        update_comment_and_inspection_columns(output_file)

        # Calculate processing time
        end_time = time.time()
        duration = end_time - start_time

        # Create full results dictionary
        result_dict = {
            'status': 'success',
            'input_file': input_file,
            'output_file': output_file,
            'processing_time': duration,
            'total_material_po_combinations': len(material_po_pairs),
            'factory_orders_found': len(factory_order_data),
            'invoice_records_found': len([k for k, v in invoice_results.items() if not v.empty]),
            'shipped_qty_column_added': shipped_qty_added,
            'comment_inspection_updated': True,
            'message': f'Successfully processed {input_file} in {duration:.2f} seconds'
        }

        print(f"\n{'='*80}")
        print(f"✅ PROCESSING COMPLETED SUCCESSFULLY!")
        print(f"📁 Output file: {output_file}")
        print(f"⏱️ Total processing time: {duration:.2f} seconds")
        print(f"{'='*80}")

        # Return based on simple_return parameter
        if simple_return:
            return output_file
        else:
            return result_dict

    except Exception as e:
        error_result = {
            'status': 'error',
            'input_file': input_file,
            'output_file': output_file if 'output_file' in locals() else None,
            'error': str(e),
            'message': f'Error processing {input_file}: {str(e)}'
        }

        if simple_return:
            return None
        else:
            return error_result

